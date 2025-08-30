"""
EU MRV Emissions Analytics
==========================

This script processes the official EMSA EU MRV Excel files (2018–2024)
to compute annual total CO₂ emissions across all ships. It produces:

1. Annual totals per year (2018–2024).
2. A CSV file with results saved to /results/EU_MRV_CO2_Totals.csv.
3. A line chart of annual CO₂ trends saved as /results/EU_MRV_CO2_Trends.png.

Author: Mahtab Shahin
Date: 2025
"""

import os
import csv
import matplotlib.pyplot as plt
from openpyxl import load_workbook

# --- Configuration ---
DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
RESULTS_DIR = os.path.join(os.path.dirname(__file__), "..", "results")
os.makedirs(RESULTS_DIR, exist_ok=True)

# Mapping year -> filename
files = {
    2018: "2018-v272-05072025-EU MRV Publication of information.xlsx",
    2019: "2019-v227-28052025-EU MRV Publication of information.xlsx",
    2020: "2020-v207-23082025-EU MRV Publication of information.xlsx",
    2021: "2021-v214-30042025-EU MRV Publication of information.xlsx",
    2022: "2022-v238-01082025-EU MRV Publication of information.xlsx",
    2023: "2023-v69-29082025-EU MRV Publication of information.xlsx",
    2024: "2024-v55-30082025-EU MRV Publication of information.xlsx",
}


def sum_co2_excel(path, sheet_name, col_name="Total CO₂ emissions [m tonnes]", header_row=3):
    """
    Efficiently sum CO₂ emissions from an EMSA MRV Excel file.

    Args:
        path (str): Path to Excel file.
        sheet_name (str): Sheet containing annual data.
        col_name (str): Target column header.
        header_row (int): Row number where headers are located (default = 3).

    Returns:
        float: Total CO₂ emissions in tonnes.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Locate column index
    headers = [cell.value for cell in ws[header_row]]
    try:
        col_idx = headers.index(col_name) + 1  # convert to 1-based index
    except ValueError:
        print(f"[ERROR] Column '{col_name}' not found in {path}")
        wb.close()
        return None

    total = 0.0
    for row in ws.iter_rows(min_row=header_row + 1,
                            min_col=col_idx, max_col=col_idx,
                            values_only=True):
        val = row[0]
        if isinstance(val, (int, float)):
            total += val

    wb.close()
    return total


def main():
    results = []
    for year, filename in files.items():
        path = os.path.join(DATA_DIR, filename)
        sheet = "2024 Full ERs" if year == 2024 else str(year)

        print(f"Processing {year} ...")
        total = sum_co2_excel(path, sheet)

        if total is not None:
            total_mt = total / 1e6  # convert tonnes -> million tonnes
            results.append((year, total_mt))
            print(f"  ✔ {year}: {total_mt:.2f} Mt CO₂")
        else:
            print(f"  ✘ Failed to read {year}")

    # Save results as CSV
    csv_path = os.path.join(RESULTS_DIR, "EU_MRV_CO2_Totals.csv")
    with open(csv_path, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["Year", "Total_CO2_Mt"])
        writer.writerows(results)

    # Plot results
    years, totals = zip(*results)
    plt.figure(figsize=(8, 5))
    plt.plot(years, totals, marker="o", linewidth=2, color="tab:blue")
    plt.title("Annual Total CO₂ Emissions in EU MRV (2018–2024)", fontsize=13)
    plt.xlabel("Year")
    plt.ylabel("CO₂ Emissions [million tonnes]")
    plt.grid(True, linestyle="--", alpha=0.6)
    plt.tight_layout()
    plot_path = os.path.join(RESULTS_DIR, "EU_MRV_CO2_Trends.png")
    plt.savefig(plot_path)
    plt.close()

    print(f"\nResults saved to:\n  - {csv_path}\n  - {plot_path}")


if __name__ == "__main__":
    main()
